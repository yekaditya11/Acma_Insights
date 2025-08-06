import openpyxl
import csv
import re
import logging
from pathlib import Path

# Configure logging
logger = logging.getLogger(__name__)

def normalize_sheet_name(sheet_name: str) -> str:
    """Consistent sheet name normalization used across the application"""
    clean_name = re.sub(r'[^\w\-_]', '_', sheet_name.strip())
    clean_name = re.sub(r'_+', '_', clean_name).strip('_')
    return clean_name

def get_sheet_names(file_path):
    """Get all sheet names from Excel file"""
    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        names = workbook.sheetnames
        workbook.close()
        logger.info(f"Found {len(names)} sheets: {names}")
        return names
    except Exception as e:
        logger.error(f"Failed to load sheet names: {e}")
        return []

def is_empty_row(row):
    """Check if a row is empty"""
    return all(cell is None or str(cell).strip() == '' for cell in row)

def has_meaningful_content(row):
    """Check if a row has meaningful content"""
    non_empty_cells = [cell for cell in row if cell is not None and str(cell).strip()]
    if len(non_empty_cells) < 2:
        return False
    formula_only = all(str(cell).startswith('=') for cell in non_empty_cells)
    return not (formula_only and len(non_empty_cells) < 3)

def find_data_boundaries(sheet, start_row=6):
    """Find the boundaries of meaningful data in a sheet"""
    rows = list(sheet.iter_rows(min_row=start_row, values_only=True))
    last_meaningful_row = -1
    for i, row in enumerate(rows):
        if has_meaningful_content(row):
            last_meaningful_row = i
    return rows[:last_meaningful_row + 4] if last_meaningful_row >= 0 else rows[:10]

def extract_csv(file_path, output_dir, sheets_to_process=None, skip_first_sheet=True):
    """Extract CSV files from Excel sheets"""
    all_sheet_names = get_sheet_names(file_path)
    if not all_sheet_names:
        logger.error("No sheets found in Excel file")
        return [], {}

    target_sheets = sheets_to_process if sheets_to_process else (
        all_sheet_names[1:] if skip_first_sheet else all_sheet_names
    )

    logger.info(f"Processing {len(target_sheets)} sheet(s): {target_sheets}")

    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    csv_paths = []
    name_mapping = {}

    # Create a mapping of exact sheet names for better matching
    exact_sheet_names = {sheet.strip(): sheet for sheet in workbook.sheetnames}
    
    for sheet_name in target_sheets:
        try:
            logger.info(f"Processing sheet: '{sheet_name}'")
            
            # Try to find the exact sheet in workbook
            actual_sheet_name = None
            if sheet_name in workbook.sheetnames:
                actual_sheet_name = sheet_name
            elif sheet_name.strip() in exact_sheet_names:
                actual_sheet_name = exact_sheet_names[sheet_name.strip()]
            else:
                # Try to find by stripped comparison
                for wb_sheet in workbook.sheetnames:
                    if wb_sheet.strip() == sheet_name.strip():
                        actual_sheet_name = wb_sheet
                        break
            
            if not actual_sheet_name:
                logger.error(f"Sheet '{sheet_name}' not found in workbook")
                logger.error(f"Available sheets: {workbook.sheetnames}")
                continue
                
            logger.info(f"Using actual sheet name: '{actual_sheet_name}'")
            sheet = workbook[actual_sheet_name]
            
            rows = find_data_boundaries(sheet, start_row=6)
            non_empty_rows = [row for row in rows if not is_empty_row(row)]
            
            if not non_empty_rows:
                logger.warning(f"No meaningful content found in sheet: {sheet_name}")
                # Try to find meaningful content starting from row 1
                rows_from_start = find_data_boundaries(sheet, start_row=1)
                non_empty_rows_from_start = [row for row in rows_from_start if not is_empty_row(row)]
                if not non_empty_rows_from_start:
                    logger.warning(f"Retrying from row 1: found {len(non_empty_rows_from_start)} non-empty rows")
                    continue
                else:
                    rows = rows_from_start
                    non_empty_rows = non_empty_rows_from_start

            max_cols = max(len(row) for row in non_empty_rows)
            clean_name = normalize_sheet_name(sheet_name)
            csv_path = output_dir / f"{clean_name}.csv"

            with open(csv_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                for row in rows:
                    padded_row = list(row)
                    if len(padded_row) < max_cols:
                        padded_row.extend([None] * (max_cols - len(padded_row)))

                    cleaned = []
                    for i, cell in enumerate(padded_row):
                        if cell is None:
                            cleaned.append('')
                        elif isinstance(cell, str) and cell.strip() == '':
                            cleaned.append('')
                        elif cell == 0 and i >= 4: 
 
                            if i == 4:  
                                cleaned.append('')  # Treat as empty
                            else:
                                cleaned.append(str(cell).strip())  # Keep as 0
                        else:
                            cleaned.append(str(cell).strip())

                    writer.writerow(cleaned)

            csv_paths.append(csv_path)
            name_mapping[clean_name] = actual_sheet_name  # Use the actual sheet name found
            logger.info(f"Created CSV: {csv_path}")
            
        except KeyError as ke:
            logger.error(f"Sheet '{sheet_name}' not found in workbook")
            logger.error(f"Available sheets: {workbook.sheetnames}")
            # Try to find a close match
            available_sheets = workbook.sheetnames
            close_matches = [s for s in available_sheets if sheet_name.strip() in s or s in sheet_name.strip()]
            if close_matches:
                logger.info(f"Possible matches: {close_matches}")
            continue
        except Exception as e:
            logger.error(f"Failed to process sheet {sheet_name}: {e}")
            continue

    workbook.close()
    logger.info(f"Successfully saved {len(csv_paths)} CSV files")
    
    if not csv_paths:
        logger.warning("No CSV files were generated. This could be due to:")
        logger.warning("- All sheets have no meaningful content")
        logger.warning("- All target sheets were excluded or not found")
        logger.warning("- Processing errors occurred for all sheets")
    
    return csv_paths, name_mapping