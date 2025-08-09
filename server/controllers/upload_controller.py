from fastapi import APIRouter, UploadFile, File, HTTPException
import json
import logging
import openpyxl
import io

from config import CSV_DIR, RESULTS_DIR
from services.csv_parser import extract_csv, normalize_sheet_name
from services.kpi_builder import build_kpi_json
from services.general_summary_service import generate_general_insights
from services.kpi_ingest_service import ingest_final_kpis, test_db_connection

logger = logging.getLogger(__name__)

router = APIRouter()


@router.post("/upload_excel/")
async def upload_excel(file: UploadFile = File(...)):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported.")

    # Load workbook directly from uploaded bytes (writes a temp XLSX only for CSV extraction)
    file_bytes = await file.read()
    workbook = openpyxl.load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)

    try:
        all_sheet_names = workbook.sheetnames
        if not all_sheet_names:
            raise HTTPException(status_code=400, detail="No sheets found in the Excel file")

        from config import EXCLUDED_SHEETS
        sheets_to_process = [sheet for sheet in all_sheet_names if sheet.strip() not in EXCLUDED_SHEETS]
        if not sheets_to_process:
            raise HTTPException(status_code=400, detail="No processable sheets found after exclusions")

        sheets_with_content = []
        for sheet_name in sheets_to_process:
            try:
                if sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    rows = list(sheet.iter_rows(min_row=1, max_row=20, values_only=True))
                    non_empty_rows = [row for row in rows if any(cell is not None and str(cell).strip() for cell in row)]
                    if len(non_empty_rows) >= 2:
                        sheets_with_content.append(sheet_name)
                    else:
                        logger.warning(f"Sheet '{sheet_name}' has no meaningful content - skipping")
            except Exception as e:
                logger.warning(f"Error checking sheet '{sheet_name}': {e}")
                sheets_with_content.append(sheet_name)
        workbook.close()

        existing_csv_files = []
        missing_csv_files = []
        for sheet in sheets_to_process:
            csv_path = CSV_DIR / (normalize_sheet_name(sheet) + ".csv")
            if csv_path.exists():
                existing_csv_files.append(sheet)
            elif sheet in sheets_with_content:
                missing_csv_files.append(sheet)

        if missing_csv_files:
            # Re-parse from in-memory workbook by saving temp CSVs directly
            tmp_path = CSV_DIR  # write CSV outputs only
            tmp_path.mkdir(parents=True, exist_ok=True)
            # Use existing extract_csv with a temporary file fallback
            # Save a temp XLSX only for parsing to CSVs
            temp_xlsx = RESULTS_DIR / "temp_upload.xlsx"
            with open(temp_xlsx, "wb") as f:
                f.write(file_bytes)
            try:
                extract_csv(str(temp_xlsx), tmp_path, sheets_to_process=missing_csv_files, skip_first_sheet=False)
            finally:
                try:
                    temp_xlsx.unlink()
                except Exception:
                    pass

        all_csv_paths = []
        for sheet in sheets_to_process:
            csv_path = CSV_DIR / (normalize_sheet_name(sheet) + ".csv")
            if csv_path.exists():
                all_csv_paths.append(csv_path)

        if not all_csv_paths:
            raise HTTPException(status_code=400, detail="No CSV files available for processing")

        supplier_kpi_info = build_kpi_json()
        if supplier_kpi_info is None:
            logger.warning("KPI builder returned None; continuing with empty outputs")
            supplier_kpi_info = {}

        # Generate general insights first (sync)
        try:
            general = generate_general_insights()
        except Exception as e:
            logger.warning(f"General insights generation failed: {e}")
            general = []

        ingest_result = {"upserted": 0}
        try:
            # quick connectivity check to fail fast
            if test_db_connection():
                ingest_result = ingest_final_kpis(str(RESULTS_DIR / "final_supplier_kpis.json"), skip_nulls=False)
            else:
                logger.warning("Skipping ingestion: DB connectivity test failed")
        except Exception as ingest_err:
            logger.warning(f"KPI ingestion failed: {ingest_err}")

        return {
            "message": "Processing completed",
            "general-insights": general,
            "Supplier-KPIs": supplier_kpi_info,
            "ingestion": ingest_result,
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Unexpected error during processing")
        raise HTTPException(status_code=500, detail=f"Processing failed: {str(e)}")


