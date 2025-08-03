from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import RedirectResponse
from fastapi.middleware.cors import CORSMiddleware
from pathlib import Path
import shutil
import os
import json
import logging
import openpyxl

from sheet_insights.parser import extract_csv, get_sheet_names, normalize_sheet_name
from sheet_insights.insights import get_insights
from sheet_insights.general_summary import generate_general_insights
from sheet_insights.kpi_dashboard import get_all_supplier_kpi_json
from sheet_insights.additional_insights import generate_additional_insights
from config import (
    HOST, PORT, DEBUG, UPLOAD_DIR, CSV_DIR, INSIGHTS_FILE, 
    RESULTS_DIR, OUTPUT_JSON, ALLOWED_ORIGINS, LOG_LEVEL, LOG_FORMAT
)

# Configure logging
logging.basicConfig(
    level=getattr(logging, LOG_LEVEL),
    format=LOG_FORMAT
)
logger = logging.getLogger(__name__)

app = FastAPI(
    title="Excel Parser API",
    description="API for processing Excel files and generating insights",
    version="1.0.0"
)

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Create necessary directories
for folder in [UPLOAD_DIR, CSV_DIR, RESULTS_DIR]:
    folder.mkdir(parents=True, exist_ok=True)

def normalize_filename(sheet_name: str) -> str:
    """Use consistent normalization with parser.py"""
    return normalize_sheet_name(sheet_name) + ".csv"

@app.get("/")
def read_root():
    """Redirect to API documentation"""
    return RedirectResponse(url='/docs')

@app.get("/health")
def health_check():
    """Health check endpoint"""
    return {
        "status": "healthy",
        "service": "Excel Parser API",
        "version": "1.0.0"
    }

@app.get("/sheet_insights")
def get_sheet_insights():
    """Get individual sheet insights for deep dive"""
    try:
        insights_file = Path('results/insights.json')
        if not insights_file.exists():
            raise HTTPException(
                status_code=404, 
                detail="Insights file not found. Please upload and process an Excel file first."
            )
        
        with open(insights_file, "r", encoding="utf-8") as f:
            insights = json.load(f)
        
        return {"insights": insights}
    except Exception as e:
        logger.error(f"Error getting sheet insights: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to get sheet insights: {str(e)}")

@app.post("/generate_more_insights")
def generate_more_insights():
    """Generate additional insights from the data"""
    try:
        # Check if KPI data exists
        kpi_file = Path('results/final_supplier_kpis.json')
        if not kpi_file.exists():
            raise HTTPException(
                status_code=404, 
                detail="KPI data not found. Please upload and process an Excel file first."
            )
        
        logger.info("Generating additional insights...")
        
        # Generate new additional insights
        additional_insights = generate_additional_insights()
        
        # Regenerate sheet insights for fresh perspective
        logger.info("Regenerating sheet insights...")
        sheet_insights = get_insights()
        
        # Save the new sheet insights
        with open(INSIGHTS_FILE, "w", encoding="utf-8") as f:
            json.dump(sheet_insights, f, indent=2, ensure_ascii=False)
        
        # Load existing general insights
        general_file = Path('results/General-info.json')
        existing_general = []
        if general_file.exists():
            with open(general_file, "r", encoding="utf-8") as f:
                existing_general = json.load(f)
        
        # Save additional insights to a separate file
        additional_file = Path('results/additional-insights.json')
        with open(additional_file, "w", encoding="utf-8") as f:
            json.dump(additional_insights, f, indent=2, ensure_ascii=False)
        
        logger.info("Additional insights generated successfully")
        
        return {
            "message": "Additional insights generated successfully",
            "additional_insights": additional_insights,
            "existing_general_insights": existing_general,
            "sheet_insights": sheet_insights,
            "total_additional_insights": len(additional_insights)
        }
        
    except Exception as e:
        logger.error(f"Error generating more insights: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to generate more insights: {str(e)}")

@app.post("/upload_excel/")
async def upload_excel(file: UploadFile = File(...)):
    """Upload and process Excel file"""
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported.")

    file_path = UPLOAD_DIR / file.filename
    with open(file_path, "wb") as f:
        shutil.copyfileobj(file.file, f)

    logger.info(f"Uploaded file saved: {file_path.name}")

    try:
        all_sheet_names = get_sheet_names(str(file_path))
        if not all_sheet_names:
            raise HTTPException(status_code=400, detail="No sheets found in the Excel file")

        logger.info(f"All sheets found in Excel: {all_sheet_names}")

        # Define excluded sheets
        from config import EXCLUDED_SHEETS
        sheets_to_process = [sheet for sheet in all_sheet_names if sheet.strip() not in EXCLUDED_SHEETS]

        logger.info(f"Sheets selected for processing: {sheets_to_process}")
        
        # Check if we have any sheets to process
        if not sheets_to_process:
            logger.warning("No sheets available for processing after exclusions")
            logger.warning(f"All sheets: {all_sheet_names}")
            logger.warning(f"Excluded: {EXCLUDED_SHEETS}")
            raise HTTPException(
                status_code=400, 
                detail=f"No processable sheets found. Available sheets: {all_sheet_names}. Excluded: {EXCLUDED_SHEETS}"
            )

        # Check for existing CSV files
        existing_csv_files = []
        missing_csv_files = []
        
        # Check which sheets have meaningful content
        workbook = openpyxl.load_workbook(str(file_path), read_only=True)
        sheets_with_content = []
        
        for sheet_name in sheets_to_process:
            try:
                if sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    # Check if sheet has meaningful content
                    rows = list(sheet.iter_rows(min_row=1, max_row=20, values_only=True))
                    non_empty_rows = [row for row in rows if any(cell is not None and str(cell).strip() for cell in row)]
                    if len(non_empty_rows) >= 2:  # At least 2 rows with content
                        sheets_with_content.append(sheet_name)
                    else:
                        logger.warning(f"Sheet '{sheet_name}' has no meaningful content - skipping")
            except Exception as e:
                logger.warning(f"Error checking sheet '{sheet_name}': {e}")
                # Assume it has content if we can't check
                sheets_with_content.append(sheet_name)
        
        workbook.close()
        
        for sheet in sheets_to_process:
            csv_path = CSV_DIR / normalize_filename(sheet)
            if csv_path.exists():
                existing_csv_files.append(sheet)
                logger.info(f"Found existing CSV: {csv_path}")
            elif sheet in sheets_with_content:
                missing_csv_files.append(sheet)
                logger.info(f"Missing CSV: {csv_path}")
            else:
                logger.info(f"Skipping empty sheet: {sheet}")

        logger.info(f"Existing CSV files: {len(existing_csv_files)}")
        logger.info(f"Missing CSV files: {len(missing_csv_files)}")

        # Generate missing CSV files
        if missing_csv_files:
            logger.info(f"Generating CSV files for sheets: {missing_csv_files}")
            csv_paths, name_mapping = extract_csv(
                str(file_path),
                CSV_DIR,
                sheets_to_process=missing_csv_files,
                skip_first_sheet=False,
            )
            
            # Verify CSV files were actually created
            actual_csv_paths = [path for path in csv_paths if path.exists()]
            logger.info(f"Successfully generated CSV files: {[p.name for p in actual_csv_paths]}")
            
            # Don't fail if no CSV files were generated - this can happen with empty sheets
            if len(actual_csv_paths) == 0:
                logger.warning("No CSV files were generated. This could be due to sheets having no meaningful content.")
                logger.warning("Continuing with existing CSV files...")
        else:
            logger.info("All CSV files already exist. Skipping generation.")
            actual_csv_paths = []
            name_mapping = {}

        # Collect all available CSV files
        all_csv_paths = []
        for sheet in sheets_to_process:
            csv_path = CSV_DIR / normalize_filename(sheet)
            if csv_path.exists():
                all_csv_paths.append(csv_path)

        logger.info(f"Total CSV files available for processing: {len(all_csv_paths)}")
        logger.info(f"CSV files: {[p.name for p in all_csv_paths]}")

        if not all_csv_paths:
            # Provide detailed error information
            error_details = {
                "total_sheets": len(all_sheet_names),
                "excluded_sheets": EXCLUDED_SHEETS,
                "sheets_to_process": sheets_to_process,
                "csv_directory": str(CSV_DIR),
                "expected_files": [normalize_filename(sheet) for sheet in sheets_to_process]
            }
            logger.error(f"Detailed error info: {error_details}")
            
            raise HTTPException(
                status_code=400, 
                detail=f"No CSV files available for processing. Details: {error_details}"
            )

        # Process KPIs and insights
        logger.info("Generating supplier KPI data...")
        supplier_kpi_info = get_all_supplier_kpi_json()
        logger.info("Created final_supplier_kpis.json")

        logger.info("Generating insights...")
        insights = get_insights()
        with open(INSIGHTS_FILE, "w", encoding='utf-8') as f:
            json.dump(insights, f, indent=2, ensure_ascii=False)
        logger.info(f"Saved insights to: {INSIGHTS_FILE}")

        logger.info("Generating general insights...")
        general = generate_general_insights()

        # Load insights for response
        with open(INSIGHTS_FILE, "r", encoding="utf-8") as f:
            insights_content = json.load(f)

        logger.info("Processing completed successfully.")

        return {
            "insights": insights_content,
            "general-insights": general,
            "Supplier-KPIs": supplier_kpi_info
        }

    except HTTPException:
        # Re-raise HTTP exceptions as-is
        raise
    except Exception as e:
        logger.error(f"Unexpected error during processing: {e}")
        logger.error(f"Error type: {type(e).__name__}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Processing failed: {str(e)}")

if __name__ == "__main__":
    import uvicorn

    logger.info(f"Starting FastAPI server on {HOST}:{PORT}...")
    uvicorn.run(
        "app:app",
        host=HOST,
        port=PORT,
        reload=DEBUG,
        log_level="info"
    )